import os
import pdfplumber
import pandas as pd
from openpyxl import load_workbook

def process_pdf_file(pdf_path, final_path):

    temp_excel = "temp_output.xlsx"

    # ===============================
    # Step 1 PDF 转 Excel
    # ===============================
    with pdfplumber.open(pdf_path) as pdf:
        writer = pd.ExcelWriter(temp_excel, engine='openpyxl')
        table_count = 0

        for page_number, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()

            for table in tables:
                df = pd.DataFrame(table)
                sheet_name = f"Page{page_number}_Table{table_count+1}"
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                table_count += 1

        writer.close()

    print("PDF extraction finished.")

    # ===============================
    # Step 2 读取 Excel（⚠️ 唯一适配修改：自动找sheet）
    # ===============================
    all_sheets = pd.read_excel(temp_excel, sheet_name=None, header=None)
    sheet_names = list(all_sheets.keys())

    if len(sheet_names) < 2:
        raise Exception("Not enough tables found in PDF")

    df1 = all_sheets[sheet_names[0]]
    df2 = all_sheets[sheet_names[1]]

    wb = load_workbook(temp_excel)

    for sheet in ["NS", "DS"]:
        if sheet in wb.sheetnames:
            del wb[sheet]

    ws_ns = wb.create_sheet("NS")
    ws_ds = wb.create_sheet("DS")

    # ===============================
    # 列索引（完全保留）
    # ===============================
    col_D = 3
    col_E = 4
    col_F = 5
    col_G = 6
    col_I = 8
    col_J = 9
    col_L = 11
    col_N = 13
    col_Q = 16

    # ===============================
    # 清理转数字函数（完全保留）
    # ===============================
    def clean_and_convert(val):
        if pd.isna(val) or val == "":
            return None
        try:
            return float(str(val).replace(" ", "").replace(",", ""))
        except:
            return val

    # ===============================
    # 标题（完全保留）
    # ===============================
    ws_ns.append([
        df1.iloc[8, col_D],
        df1.iloc[8, col_I],
        df1.iloc[8, col_N]
    ])

    ws_ds.append([
        df2.iloc[9, col_D],
        df2.iloc[9, col_I],
        df2.iloc[9, col_N]
    ])

    # ===============================
    # NS 数据（完全保留）
    # ===============================
    for r in range(10, 65):

        if pd.notna(df1.iloc[r, col_L]) and df1.iloc[r, col_L] != 0:
            ws_ns.append([
                df1.iloc[r, col_D],
                df1.iloc[r, col_E],
                df1.iloc[r, col_F],
                df1.iloc[r, col_G],
                df1.iloc[r, col_J],
                df1.iloc[r, col_L],
                df1.iloc[r, col_Q]
            ])

        if pd.notna(df1.iloc[r, col_N]) and df1.iloc[r, col_N] != 0:
            ws_ns.append([
                df1.iloc[r, col_D],
                df1.iloc[r, col_E],
                df1.iloc[r, col_F],
                df1.iloc[r, col_G],
                df1.iloc[r, col_J],
                df1.iloc[r, col_N],
                df1.iloc[r, col_Q]
            ])

    # ===============================
    # DS 数据（完全保留）
    # ===============================
    for r in range(11, 66):

        if pd.notna(df2.iloc[r, col_L]) and df2.iloc[r, col_L] != 0:
            ws_ds.append([
                df2.iloc[r, col_D],
                df2.iloc[r, col_E],
                df2.iloc[r, col_F],
                df2.iloc[r, col_G],
                df2.iloc[r, col_J],
                df2.iloc[r, col_L],
                df2.iloc[r, col_Q]
            ])

        if pd.notna(df2.iloc[r, col_N]) and df2.iloc[r, col_N] != 0:
            ws_ds.append([
                df2.iloc[r, col_D],
                df2.iloc[r, col_E],
                df2.iloc[r, col_F],
                df2.iloc[r, col_G],
                df2.iloc[r, col_J],
                df2.iloc[r, col_N],
                df2.iloc[r, col_Q]
            ])

    # ===============================
    # 数据转数字（完全保留！！）
    # ===============================
    def convert_sheet(ws):
        max_row = ws.max_row

        for col in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col).value = clean_and_convert(
                ws.cell(row=1, column=col).value
            )

        for r in range(1, max_row + 1):
            ws.cell(row=r, column=6).value = clean_and_convert(
                ws.cell(row=r, column=6).value
            )
            ws.cell(row=r, column=7).value = clean_and_convert(
                ws.cell(row=r, column=7).value
            )

    convert_sheet(ws_ns)
    convert_sheet(ws_ds)

    # ===============================
    # 列剪切重排（完全保留！！）
    # ===============================
    def rearrange_sheet(ws):

        max_row = ws.max_row

        a1 = ws["A1"].value
        b1 = ws["B1"].value
        c1 = ws["C1"].value

        ws["L1"] = a1
        ws["M1"] = b1
        ws["N1"] = c1

        ws["A1"] = None
        ws["B1"] = None
        ws["C1"] = None

        col_A = [ws.cell(r, 1).value for r in range(1, max_row + 1)]
        for r in range(1, max_row + 1):
            ws.cell(r, 11).value = col_A[r - 1]
            ws.cell(r, 1).value = None

        col_D = [ws.cell(r, 4).value for r in range(1, max_row + 1)]
        for r in range(1, max_row + 1):
            ws.cell(r, 1).value = col_D[r - 1]
            ws.cell(r, 4).value = None

        col_B = [ws.cell(r, 2).value for r in range(1, max_row + 1)]
        for r in range(1, max_row + 1):
            ws.cell(r, 4).value = col_B[r - 1]
            ws.cell(r, 2).value = None

        col_E = [ws.cell(r, 5).value for r in range(1, max_row + 1)]
        for r in range(1, max_row + 1):
            ws.cell(r, 10).value = col_E[r - 1]
            ws.cell(r, 5).value = None

        col_C = [ws.cell(r, 3).value for r in range(1, max_row + 1)]
        for r in range(1, max_row + 1):
            ws.cell(r, 5).value = col_C[r - 1]
            ws.cell(r, 3).value = None

        col_F = [ws.cell(r, 6).value for r in range(1, max_row + 1)]
        for r in range(1, max_row + 1):
            ws.cell(r, 9).value = col_F[r - 1]
            ws.cell(r, 6).value = None

        for r in range(2, max_row + 1):
            val = ws.cell(r, 11).value
            if val:
                val_str = str(val)
                if "-" in val_str:
                    left, right = val_str.split("-", 1)
                    ws.cell(r, 2).value = left
                    ws.cell(r, 3).value = right
                else:
                    ws.cell(r, 3).value = val_str

        ws["B1"] = ws["L1"].value
        ws["C1"] = ws["M1"].value
        ws["A1"] = ws["N1"].value

        ws["L1"] = None
        ws["M1"] = None
        ws["N1"] = None

    rearrange_sheet(ws_ns)
    rearrange_sheet(ws_ds)

    # ===============================
    # 删除列（完全保留）
    # ===============================
    ws_ns.delete_cols(11)
    ws_ns.delete_cols(10)

    ws_ds.delete_cols(11)
    ws_ds.delete_cols(10)

    # ===============================
    # 保存
    # ===============================
    wb.save(final_path)
    os.remove(temp_excel)

    print("All processing complete. Final file saved to:")
    print(final_path)
